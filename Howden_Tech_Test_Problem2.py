#!/usr/bin/env python

import requests
import pandas as pd
import json

import sys, fnmatch, os, subprocess, copy, getpass
import openpyxl
from datetime import datetime #, timedelta
#import timedelta

from openpyxl import load_workbook, workbook
from openpyxl.styles import Border, Side, Alignment, Color, colors, PatternFill, Font, Border, Protection, NamedStyle
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from openpyxl.chart import PieChart3D, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import PieChart, ProjectedPieChart, Reference, BarChart, Series
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.text import RichText
from openpyxl.chart.plotarea import DataTable
from copy import deepcopy


arg_target_path = sys.argv[1]

print("\n")
print("A warm welcome to the Howden Tech Test Python Program for Problem2 - Exchange Rates API.")
print("After the program is run, see the output in the Workbook created in location:")
print(arg_target_path)


os.chdir(arg_target_path)

DateTimeStr = datetime.now().strftime("%d$b%Y_%H%M%S_%f")
filename = 'Howden_Test_Result2_'+DateTimeStr+'.xlsx'
filepath = arg_target_path + filename
workbook = openpyxl.Workbook()
workbook.save(filepath)
workbook.close()
wb = load_workbook(filename)

ws_exchange = wb.create_sheet('ExchangeRates')
ws_exchange = wb["ExchangeRates"]

ws_exchange_dummy = wb.create_sheet('Dummy')
ws_exchange_dummy = wb["Dummy"]

###################################################################################################################

def adjust_column_with_from_col(ws, min_row, min_col, max_col):
    column_widths = []
    for i, col in \
        enumerate(
            ws.iter_cols(min_col = min_col, max_col = max_col, min_row = min_row)
        ):
        for cell in col:
            value = cell.value
            if(value is not None):
                if(isinstance(value, str) is False):
                    value = str(value)
                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))
                    
    for i, width in enumerate(column_widths):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        ws.column_dimensions[col_name].width = value
        
###################################################################################################################

url = 'https://v6.exchangerate-api.com/v6/9e9c79edd6ab5b26d2a49522/latest/USD'

response = requests.get(url)
data = response.json()
#print(data)

json_filename = 'ExchangeRates_'+DateTimeStr+'.json'

with open(json_filename, 'w') as json_file:
    json.dump(data, json_file)

df = pd.read_json(json_filename)
#print(df)

rows_exchange_rates = dataframe_to_rows(df)
for i, line in enumerate(rows_exchange_rates):
    for k, val in enumerate(line):
        ws_exchange_dummy.cell(i+1, k+1).value = val

ws_exchange_dummy.delete_rows(idx=1, amount=2)
ws_exchange_dummy.delete_cols(idx=2, amount=8)

wb.save(filename)

required_curr_list = ['AUD', 'CAD', 'CHF', 'CNY', 'EUR', 'GBP', 'HKD', 'JPY', 'NZD', 'USD']

ws_exchange.cell(1, 1).value = "Rate Type"
ws_exchange.cell(1, 2).value = "Date"
ws_exchange.cell(1, 3).value = "Currency_From"
ws_exchange.cell(1, 4).value = "Currency_From_Value"
ws_exchange.cell(1, 5).value = "Currency_To"
ws_exchange.cell(1, 6).value = "Currency_To_Value"

for i in range(2, 12):
    ws_exchange.cell(i, 1).value = "Spot rate"
    ws_exchange.cell(i, 2).value = str(datetime.now().strftime("%d/%m/%Y"))
    ws_exchange.cell(i, 3).value = "USD"
    ws_exchange.cell(i, 4).value = "1"
    
k = 2
row_range = ws_exchange_dummy.max_row
for i in range(1, row_range+1):
    for j in range(0, len(required_curr_list)):
        if(ws_exchange_dummy.cell(i, 1).value == required_curr_list[j]):
            ws_exchange.cell(k, 5).value = ws_exchange_dummy.cell(i, 1).value
            ws_exchange.cell(k, 6).value = ws_exchange_dummy.cell(i, 2).value
            k = k + 1
            continue

wb.save(filename)

thin = Side(border_style = "thin")

for row in ws_exchange.iter_rows():
    for cell in row:
        cell.border = Border(top = thin, left = thin, right = thin, bottom = thin)

adjust_column_with_from_col =(ws_exchange, 1, 1, ws_exchange.max_column)

col_range = ws_exchange.max_column
row_range = ws_exchange.max_row
for col in range(1, col_range + 1):
    cell_header = ws_exchange.cell(1, col)
    cell_header.fill = PatternFill(start_color = '0000ff00', end_color = '0000ff00', fill_type = 'solid')   #green colored heading
    
wb.save(filename)

#######################################################################################################################################

del wb['Dummy']
del wb['Sheet']
wb.save(filename)
wb.close()

print("\n")
print("End of Program. Good Luck - Thanks.")


    





            

