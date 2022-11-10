#!/usr/bin/env python

import sys, fnmatch, os, subprocess, copy, getpass
from datetime import datetime, timedelta
import timedelta

import cx_Oracle
import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl

from openpyxl import load_workbook, workbook
from openpyxl.styles import Border, Side, Alignment, Color, Colors, PatternFill, Font, Border, Protection, NamedStyle
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from openpyxl.chart import PieChart3D, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import PieChart, ProjectedPieChart, Reference, BarChart, Series
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.chart.text import RichText
from copy import deepcopy
from openpyxl.chart.plotarea import DataTable

try:
    if sys.platform.startswith("darwin"):
        lib_dir = os.path.join(os.environ.get("HOME"), "Downloads", "instantclient_19_12")
        cx_Oracle.init_oracle_client(lib_dir = lib_dir)
    elif sys.platform.startswith("win32"):
        lib_dir = r"C:\Oracle\instantclient_19_12")
        cx_Oracle.init_oracle_client(lib_dir = lib_dir)
except Exception as err:
    print("Whoops")
    print(err);
    sys.exit(1);


arg_source_path = sys.argv[1]
arg_target_path = sys.argv[2]

print("\n")
user_id = input("Enter the Database UserID:")
user_password = getpass.getpass(Prompt='Enter the Database Password: ' stream=None)

print("\n")
start_time = time.time()
print("Script started @: "+datetime.now().strftime("%d$b%Y_%H%M%S_%f"))

print("\n")
print("A warm welcome to the Howden Tech Test Python Program.")
print("After the program is run, see the output in the Workbook created in location:")
print(arg_target_path)
print("\n")

dsn_tns = cx_Oracle.makedsn('<Your Host Nam of DB here>', 1521, service_name = '<Your Service Name of DB here>')
conn = cx_Orcle.connect(user=user_id, password = user_password, dsn=dsn_tns, encoding = 'UTF-8')
curs = conn.cursor()

os.chdir(arg_target_path)

DateTimeStr = datetime.now().strftime("%d$b%Y_%H%M%S_%f")
filename = 'Howden_Test_Result_'+DateTimeStr+'.xlsx'
filepath = arg_target_path + filename
workbook = openpyxl.Workbook()
workbook.save(filepath)
workbook.close()
wb = load_workbook(filename)

#global ws_gl_np
ws_factstat = wb.create_sheet('FACTSTAT')
ws_factstat = wb["FACTSTAT"]

ws_factdata = wb.create_sheet('FACTDATA')
ws_factdata = WB["FACTDATA"]


###################################################################################################################

os.chdir(arg_source_path)

for file in os.listdir(arg_source_path):
    src_filename = os.path.basename(file)

class act_sheet1:
    n = 0
    src_wb = openpyxl.load_workbook(src_filename)
    sheet_names = src_wb.sheet_names
    src_sheet = sheet_names[1]
    ws1 = src_wb[src_sheet]
    gl_stat_rows = []
    for i in range(6, ws1.max_row-1):
        gl_stat_row = []
        gl_data_row = []
        for j in range(3, ws1.max_column-4):
            gl_stat_row.append(ws1.cell(i, j).value)
        gl_stat_rows.append(gl_stat_row)
    mr = ws1.max_row
    mc = ws1.max_column
    for i in range(1, mr+1):
        for j in range(1, mc+1):
            c = ws1.cell(row = i, column = j)
            ws_factdata.cell(row = i, column = j).value = c.value
    os.chdir(arg_target_path)
    wb.save(filename)
    
    
gl_stat_row_object = act_sheet1()
#print(gl_stat_row_object.gl_stat_rows)

k =0
j =0
months = 12
for i in range(2, 146):
    ws_factstat.cell(i, 5).value = gl_stat_row_object.gl_stat_rows[k][j]
    j = j + 1
    if(i == months + 1):
        i = months + 2
        months = months + 12
        k = k + 1
        j = 0
        continue
        
        
os.chdir(arg_source_path)

for file in os.listdir(arg_source_path):
    src_filename = os.path.basename(file)

class act_sheet2:
    n = 0
    src_wb = openpyxl.load_workbook(src_filename)
    sheet_names = src_wb.sheet_names
    src_sheet2 = sheet_names[2]
    ws2 = src_wb[src_sheet2]
    ma_stat_rows = []
    for i in range(6, ws2.max_row-1):
        ma_stat_row = []
        for j in range(3, ws1.max_column-4):
            ma_stat_row.append(ws2.cell(i, j).value)
        ma_stat_rows.append(ma_stat_row)
    mr = ws2.max_row
    mc = ws2.max_column
    for i in range(1, mr+1):
        for j in range(1, mc+1):
            c = ws2.cell(row = i, column = j)
            ws_factdata.cell(row = i+19, column = j).value = c.value
    os.chdir(arg_target_path)
    wb.save(filename)
    
    
ma_stat_row_object = act_sheet2()

k = 0
j = 0
months = 12
for i in range(146, 290):
    ws_factstat.cell(i, 5).value = ma_stat_row_object.ma_stat_rows[k][j]
    j = j + 1
    if(i==144+months+1):
        i = 144 + months + 2
        months = months + 12
        k = k + 1
        j = 0
        continue

#####################################################################################################

os.chdir(arg_target_path)

ws_factstat.cell(1, 1).value = "COMP_NAME"
ws_factstat.cell(1, 2).value = "LOB"
ws_factstat.cell(1, 3).value = "CURR"
ws_factstat.cell(1, 4).value = "YEAR"
ws_factstat.cell(1, 5).value = "LIR"
ws_factstat.cell(1, 6).value = "DEV_MONTH"
ws_factstat.cell(1, 7).value = "DW_CREATED_DT"
ws_factstat.cell(1, 8).value = "DW_CREATED_BY"

for i in range(2, 146):
    ws_factstat.cell(i, 1).value = "Howden"
    ws_factstat.cell(i, 2).value = "GL-np"
    ws_factstat.cell(i, 3).value = "EUR"
    ws_factstat.cell(i, 7).value = str(datetime.now().strftime("%d%b%Y"))
    ws_factstat.cell(i, 8).value = "DKB"

for i in range(146, 290):
    ws_factstat.cell(i, 1).value = "Howden"
    ws_factstat.cell(i, 2).value = "MA-np"
    ws_factstat.cell(i, 3).value = "EUR"
    ws_factstat.cell(i, 7).value = str(datetime.now().strftime("%d%b%Y"))
    ws_factstat.cell(i, 8).value = "DKB"
    

year = 2010
months = 12
for i in range(2, 146):
    ws_factstat.cell(i, 4).value = year
    if(i == months + 1):
        year = year + 1
        months = months + 12
        continue
        
year = 2010
months = 12
for i in range(146, 290):
    ws_factstat.cell(i, 4).value = year
    if(i == 144 + months + 1):
        year = year + 1
        months = months + 12
        continue
 
n = 0
dev_month = 12
for i in range(2, 146):
    ws_factstat.cell(i, 6).value = n + 12
    if(i == (n*1) + 13):
        n = n + 12
        continue 

n = 0
dev_month = 12
for i in range(146, 290):
    ws_factstat.cell(i, 6).value = n + 12
    if(i == 144 + (n*1) + 13):
        n = n + 12
        continue 


#Format FACTDATA sheet:
ws_factdata.delete_rows(idx=1, amount=4)
ws_factdata.delete_rows(idx=14, amount=7)
ws_factdata.delete_rows(idx=26, amount=2)
ws_factdata.delete_rows(idx=3, amount=13)

ws_factdata.insert_cols(1)

ws_factdata.cell(1, 1).value = "LOB"
ws_factdata.cell(1, 8).value = "ULR"
ws_factdata.cell(1, 9).value = "DWCRDT"
ws_factdata.cell(1, 10).value = "DWCRBY"

LOB = "GL-np"
for i in range(2, 14):
    ws_factdata.cell(i, 1).value = LOB
LOB = "MA-np"
for i in range(14, 26):
    ws_factdata.cell(i, 1).value = LOB
    
year = 2010
for i in range(2, 14):
    ws_factdata.cell(i, 2) = year
    year = year + 1
year = 2010
for i in range(14, 26):
    ws_factdata.cell(i, 2) = year
    year = year + 1

for i in range(2, 26):
    ws_factdata.cell(i, 8).value = ws_factdata.cell(i, 5).value + ws_factdata.cell(i, 6).value + ws_factdata.cell(i, 7).value

for i in range(2, 26):
    ws_factdata.cell(i, 9).value = datetime.now().strftime("%d%b%Y")

for i in range(2, 26):
    ws_factdata.cell(i, 10).value = "DKB"
    

del wb['Sheet']
wb.save(filename)


#################################################################################################################################

#Inserting into Corresponding tables from Excel:

gl_factstat_df = pd.read_excel(filename, sheet_name = "FACTSTAT")

gl_factstat_df['YEAR'] = gl_factstat_df['YEAR'].apply(str)
gl_factstat_df = gl_factstat_df.fillna (0)

gl_factstat_tuples = [tuple(x) for x in gl_factstat_df.values]

InsSqlText = """
INSERT INTO FACTSTAT (COMPNAME, LOBUS, CURR, YEARS, LIR, DEVMONTH, DWCRTDT, DWCRTBY) VALUES (:1, :2, :3, :4, :5, :6, :7, :8)
"""

curs.executemany(InsSqlText, gl_factstat_tuples)
conn.commit()

gl_factdata_df = pd.read_excel(filename, sheet_name = "FACTDATA")

gl_factdata_df['U/W Year'] = gl_factdata_df['U/W Year'].apply(str)
gl_factdata_df = gl_factdata_df.fillna (0)

gl_factstat_tuples = [tuple(x) for x in gl_factdata_df.values]

InsSqlText = """
INSERT INTO FACTDATA (LOBUS, YEARS, GWP, EARNEDPREM, PAIDLOSS, CASEREVERS, IBNR, ULR, DWCRTDT, DWCRTBY) VALUES (:1, :2, :3, :4, :5, :6, :7, :8, :9, :10)
"""

curs.executemany(InsSqlText, gl_factdata_tuples)
conn.commit()

####################################################################################################################################

ws_graph_ulr = wb.create_sheet('GRAPH_ULR')
ws_graph_ulr = wb["GRAPH_ULR"]

def dashboard_chart_UltimateLossRatio(filename, wb, ws_graph_ulr):
    chart_EPbyULR = BarChart()
    ws_factdata = wb["FACTDATA"]
    data = Reference(ws_factdata, min_col = 2, min_row = 1, max_row = ws_factdata.max_row, max_col = ws_factdata.max_column)
    labels = Reference(ws_factdata, min_col = 1, min_row = 2, max_row = ws_factdata.max_row)
    chart_EPbyULR.add_data(data, titles_from_data = True)
    chart_EPbyULR.plot_area.dTable = DataTable()
    chart_EPbyULR.plot_area.dTable.showHorzBorder = True
    chart_EPbyULR.plot_area.dTable.showVertBorder = True
    chart_EPbyULR.plot_area.dTable.showOutline = True
    chart_EPbyULR.plot_area.dTable.showKeys = True
    chart_EPbyULR.set_categories(labels)
    chart_EPbyULR.type = "col"
    chart_EPbyULR.style = 2
    chart_EPbyULR.grouping = "stacked"
    chart_EPbyULR.overlap = 100
    chart_EPbyULR.shape = 4
    chart_EPbyULR.title = "Ultimate Loss Ratio"
    chart_EPbyULR.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=CharacterProprties(sz=1100))
    chart_EPbyULR.height = 13.2
    chart_EPbyULR.width = 20.6
    chart_EPbyULR.legend = None
    ws_graph_ulr.add_chart(chart_EPbyULR, "D2")
    

dashboard_chart_UltimateLossRatio(filename, wb, ws_graph_ulr)


#######################################################################################################################################

wb.save(filename)
wb.close()

curs.close()
conn.close()

print("\n")
print("Script completed @: "+datetime.now().strftime("%d%b%Y_%H%M%S_%f"))
end_time = time.time()
time_taken = round((end_time-start_time)/60,2)
print("\n")
print("Total Time taken by Script: "+str(time_taken)+" minutes")

print("\n")
print("End of Program. Good Luck - Thanks.")


    





            

